# ==============================================================================
# The Pythonic Blacksmith v1.1 (The Reforged Blade)
#
# Mission: To programmatically create the PERFECTED "Buy Box Dominance Tracker"
#          Excel template, now with intelligent, consistent logic.
#
# Coded by: WesAI, Chief of Staff to the ScaleSmart Empire
# ==============================================================================

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

# --- CONFIGURATION v1.1 ---
FILENAME = "Buy_Box_Dominance_Tracker_v1.1.xlsx"
MAX_ROWS = 500  # How many rows of data the template should support

# NEW: Centralized thresholds for easy tweaking
THRESHOLDS = {
    "CRITICAL": 500,
    "AT_RISK": 100
}

# --- STYLES (Unchanged) ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, color="808080")

CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AT_RISK_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEALTHY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

def set_header_styles(sheet, max_col):
    """Applies standard header styling."""
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    sheet.freeze_panes = 'A2'

def create_instructions_sheet(wb):
    """Creates the 'Instructions' tab."""
    # This function remains the same as v1.0
    sheet = wb.active
    sheet.title = "Instructions"
    sheet['A1'] = "How to Use the Buy Box Dominance Tracker"
    sheet['A1'].font = TITLE_FONT
    instructions = [
        ("Step 1: Get the Data", "In Amazon Seller Central, navigate to: Reports > Business Reports > Detail Page Sales and Traffic by Child Item."),
        ("", "Set your desired date range (e.g., 'Last 30 Days') and download the CSV file."),
        ("Step 2: Input the Data", "Open the downloaded CSV, select all data (Ctrl+A), and copy it (Ctrl+C)."),
        ("", "Go to the 'Data_Input' tab in this workbook, click cell A1, and paste the data (Ctrl+V)."),
        ("Step 3: Analyze the Dashboard", "Go to the 'Buy_Box_Dashboard' tab. It will automatically update based on your data."),
        ("", "The dashboard prioritizes your biggest problems, highlights their status in color, and suggests actions."),
        ("Step 4: Track Your Actions", "Use the 'Action_Tracker' tab to log the promotions or changes you make and monitor their results.")
    ]
    row_num = 3
    for title, desc in instructions:
        sheet[f'A{row_num}'] = title
        sheet[f'A{row_num}'].font = Font(bold=True)
        sheet[f'B{row_num}'] = desc
        row_num += 1
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 100


def create_data_input_sheet(wb):
    """Creates the 'Data_Input' tab for raw data pasting."""
    # This function remains the same as v1.0
    sheet = wb.create_sheet("Data_Input")
    headers = ["ASIN", "SKU", "Product Name", "Sessions", "Buy Box Percentage", "(etc...)"]
    sheet.append(headers)
    set_header_styles(sheet, len(headers))
    sheet['A2'] = "DELETE THIS ROW AND PASTE YOUR RAW DATA EXPORT FROM SELLER CENTRAL HERE."
    sheet['A2'].font = NOTE_FONT

def create_dashboard_sheet(wb):
    """Creates the magical 'Buy_Box_Dashboard' tab with REFINED formulas."""
    sheet = wb.create_sheet("Buy_Box_Dashboard")

    headers = [
        "ASIN", "SKU", "Product Name", "Sessions", "Buy Box %", 
        "Priority Score", "Buy Box Status", "Suggested Action"
    ]
    sheet.append(headers)
    set_header_styles(sheet, len(headers))

    for i in range(2, MAX_ROWS + 2):
        # Data pulling formulas (adjust column letters based on actual CSV export)
        sheet[f'A{i}'] = f"=IF(ISBLANK(Data_Input!A{i}),\"\",Data_Input!A{i})" # ASIN
        sheet[f'B{i}'] = f"=IF(ISBLANK(Data_Input!B{i}),\"\",Data_Input!B{i})" # SKU
        sheet[f'C{i}'] = f"=IF(ISBLANK(Data_Input!C{i}),\"\",Data_Input!C{i})" # Product Name
        sheet[f'D{i}'] = f"=IF(ISBLANK(Data_Input!D{i}),\"\",Data_Input!D{i})" # Sessions
        sheet[f'E{i}'] = f"=IF(ISBLANK(Data_Input!E{i}),\"\",Data_Input!E{i})" # Buy Box %
        
        # Priority Score (The Brain) - Unchanged
        sheet[f'F{i}'] = f"=IF(E{i}<1, (1-E{i})*D{i}, 0)"
        
        # --- THE HOTFIX (v1.1) ---
        # The Buy Box Status is now driven by the smart Priority Score, not the dumb Buy Box %
        sheet[f'G{i}'] = f"=IF(F{i}>{THRESHOLDS['CRITICAL']}, \"Critical\", IF(F{i}>{THRESHOLDS['AT_RISK']}, \"At Risk\", \"Healthy\"))"
        
        # The Suggested Action now uses the same smart thresholds for consistency
        sheet[f'H{i}'] = f"=IF(F{i}>{THRESHOLDS['CRITICAL']}, \"High Priority: Apply Repricer/Promo NOW\", IF(F{i}>{THRESHOLDS['AT_RISK']}, \"Medium Priority: Investigate\", \"Low Priority: Monitor\"))"

    # Conditional Formatting uses the same logic
    range_str = f"G2:G{MAX_ROWS + 1}"
    sheet.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"Critical"'], fill=CRITICAL_FILL))
    sheet.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"At Risk"'], fill=AT_RISK_FILL))
    sheet.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"Healthy"'], fill=HEALTHY_FILL))

    # Column widths (Unchanged)
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 50

def create_action_tracker_sheet(wb):
    """Creates the 'Action_Tracker' tab."""
    # This function remains the same as v1.0
    sheet = wb.create_sheet("Action_Tracker")
    headers = ["Date", "ASIN", "Action Taken", "Result", "Notes"]
    sheet.append(headers)
    set_header_styles(sheet, len(headers))
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 50


def main():
    """Main function to forge the Excel workbook."""
    print("Firing up the Pythonic Blacksmith (v1.1)...")
    print(f"Reforging the Trojan Horse with superior logic: {FILENAME}")
    
    wb = openpyxl.Workbook()
    
    create_instructions_sheet(wb)
    create_data_input_sheet(wb)
    create_dashboard_sheet(wb)
    create_action_tracker_sheet(wb)
    
    wb.save(os.path.join("excel_templates", FILENAME))
    print("\nReforge complete. The weapon is now flawless.")
    print(f"The '{FILENAME}' is ready for any future deployment.")

if __name__ == "__main__":
    main()