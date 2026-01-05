# ==============================================================================
# The Pythonic Blacksmith v4.2 (The Quartermaster's Blade - Honed Edge)
#
# Mission: To forge the v1.2 "Restock Recommender" by reverse-engineering and
#          perfecting the client's core 'Days of Supply' logic, creating a
#          vastly superior, yet comfortingly familiar, system.
#
# Coded by: WesAI, Chief of Staff to the ScaleSmart Empire
# ==============================================================================

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

# --- CONFIGURATION v1.2 ---
FILENAME = "Restock_Recommender_v1.2.xlsx"
MAX_ROWS = 500
DAYS_OF_STOCK_TARGET = 60
DAYS_OF_SUPPLY_WARNING_THRESHOLD = 60 # Their sacred number

# --- STYLES ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16)
NOTE_FONT = Font(italic=True, color="808080")

ACTION_NEEDED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # Yellow (Their color!)
HEALTHY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Green

# (Functions for instructions and data input sheets remain the same as v1.1)
def set_header_styles(sheet, max_col):
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    # --- NEW: The Oracle's Control Panel ---
    sheet['O1'] = "Forecast Control Panel"
    sheet['O1'].font = Font(bold=True, size=12)
    sheet['O2'] = "Factor"
    sheet['P2'] = "Multiplier"

    factors = {
        'O3': "Upcoming Holiday:",
        'O4': "Scheduled Promotion:",
        'O5': "General Growth Trend:"
    }
    for cell, label in factors.items():
        sheet[cell] = label

    # Default Multipliers (user can change these)
    sheet['P3'] = 1.0
    sheet['P4'] = 1.0
    sheet['P5'] = 1.0

def create_instructions_sheet(wb):
    sheet = wb.active
    sheet.title = "Instructions"
    sheet['A1'] = "How to Use the Restock Recommender"
    sheet['A1'].font = TITLE_FONT
    instructions = [
        ("CRITICAL STEP 1: Download Reports", "This tool requires TWO separate reports from Seller Central."),
        ("Report A: FBA Inventory", "Go to: Reports > Fulfillment > Inventory > FBA Inventory. Click 'Request .csv Download'."),
        ("Report B: Business Report", "Go to: Reports > Business Reports > Detail Page Sales and Traffic by Child Item. Set date to 'Last 30 Days' and download."),
        ("STEP 2: Input Data", "Paste the entire contents of Report A into the 'Data_Input_Inventory' tab."),
        ("", "Paste the entire contents of Report B into the 'Data_Input_BizRpt' tab."),
        ("STEP 3: Analyze Dashboard", "Go to the 'Restock_Dashboard' tab. It will automatically calculate which products need restocking to maintain a 60-day supply.")
    ]
    row_num = 3
    for title, desc in instructions:
        sheet[f'A{row_num}'] = title
        sheet[f'A{row_num}'].font = Font(bold=True)
        sheet[f'B{row_num}'] = desc
        row_num += 1
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 100

def create_data_input_sheets(wb):
    # (Same as v1.1)
    sheet_inv = wb.create_sheet("Data_Input_Inventory")
    headers_inv = "..." # Paste full headers here
    sheet_inv.append(headers_inv.split("\t"))
    sheet_inv['A2'] = "PASTE YOUR 'FBA INVENTORY' DATA HERE"

    sheet_biz = wb.create_sheet("Data_Input_BizRpt")
    headers_biz = "..." # Paste full headers here
    sheet_biz.append(headers_biz.split("\t"))
    sheet_biz['A2'] = "PASTE YOUR 'BUSINESS REPORT' DATA HERE"

def create_dashboard_sheet(wb):
    """Creates the main Restock Dashboard with the NEW 'Days of Supply' logic."""
    sheet = wb.create_sheet("Restock_Dashboard")

    # --- NEW, IMPROVED HEADER ORDER ---
    headers = [
        "ASIN", "SKU", "Item Name", "Listing Link", "Available", "Inbound", 
        "Total Stock", "Days of Supply", "Sold (Last 30d)", "Daily Sales (Avg)", 
        f"{DAYS_OF_STOCK_TARGET}-Day Target", "RECOMMENDED UNITS", "Status"
    ]
    sheet.append(headers)
    set_header_styles(sheet, len(headers))

    for i in range(2, MAX_ROWS + 2):
        # Data Pulling (Same as v1.1)
        asin_ref = f'Data_Input_BizRpt!B{i}'
        sheet[f'A{i}'] = f'=IF(ISBLANK({asin_ref}),"",{asin_ref})'
        match_formula = f'MATCH(A{i}, Data_Input_Inventory!D:D, 0)'
        sheet[f'B{i}'] = f'=IFERROR(INDEX(Data_Input_Inventory!B:B, {match_formula}), "N/A")' # SKU
        sheet[f'C{i}'] = f'=IFERROR(INDEX(Data_Input_Inventory!E:E, {match_formula}), "N/A")' # product-name
        sheet[f'E{i}'] = f'=IFERROR(INDEX(Data_Input_Inventory!G:G, {match_formula}), 0)' # available
        sheet[f'F{i}'] = f'=IFERROR(INDEX(Data_Input_Inventory!BC:BC, {match_formula}), 0)' # inbound-quantity
        sheet[f'I{i}'] = f'=IFERROR(VLOOKUP(A{i}, Data_Input_BizRpt!B:AF, 30, FALSE), 0)' # Units Ordered
        
        # --- THE ARCHITECT'S REFINED CALCULATIONS (v1.2) ---
        sheet[f'D{i}'] = f'=IF(A{i}<>"", HYPERLINK("https://www.amazon.com/dp/"&A{i}, "View on Amazon"), "")'
        sheet[f'G{i}'] = f'=E{i}+F{i}'
        sheet[f'J{i}'] = f'=I{i}/30' # Daily Sales Average
        
        # NEW: The 'Days of Supply' Calculation with Divide-by-Zero Protection
        sheet[f'H{i}'] = f'=IFERROR(G{i}/J{i}, "")' # Total Stock / Daily Sales

        # --- THE ORACLE'S CALCULATION (v2.0) ---
        # The Daily Sales is now multiplied by our new factors
        forecasted_daily_sales = f'J{i} * $P$3 * $P$4 * $P$5'

        # The 60-Day Target is now based on this new forecast
        sheet[f'K{i}'] = f'({forecasted_daily_sales}) * {DAYS_OF_STOCK_TARGET}'

        # The Recommendation is now prescient
        sheet[f'L{i}'] = f'=IF(A{i}<>"", ROUND(MAX(0, K{i}-G{i}), 0), "")'
        sheet[f'M{i}'] = f'=IF(L{i}>0, "Restock Needed", "Healthy")'

    # --- THE UPGRADED "COLORING BOOK" ---
    # NEW: The "Honey" Logic - Highlight Days of Supply in Yellow
    sheet.conditional_formatting.add(f'H2:H{MAX_ROWS + 1}', CellIsRule(operator='lessThan', formula=[DAYS_OF_SUPPLY_WARNING_THRESHOLD], fill=WARNING_FILL))
    
    # Our Superior Logic - Highlight Recommended Units in Red
    sheet.conditional_formatting.add(f'L2:L{MAX_ROWS + 1}', CellIsRule(operator='greaterThan', formula=['0'], fill=ACTION_NEEDED_FILL))
    
    # The Status Column for clarity
    sheet.conditional_formatting.add(f'M2:M{MAX_ROWS + 1}', CellIsRule(operator='equal', formula=['"Healthy"'], fill=HEALTHY_FILL))

    # Adjust column widths
    for col_letter in ['A', 'B', 'C', 'D', 'L', 'M']:
        sheet.column_dimensions[col_letter].width = 25
        
def main():
    """Main function to forge the report."""
    print("Firing up the Pythonic Blacksmith (v4.2)...")
    print(f"Reforging the Quartermaster's Blade with the enemy's logic: {FILENAME}")
    
    wb = openpyxl.Workbook()
    
    # ... (call create_instructions and create_data_input sheets) ...
    create_instructions_sheet(wb)
    create_data_input_sheets(wb)
    create_dashboard_sheet(wb)

    wb.save(os.path.join("excel_templates", FILENAME))
    print("\nReforge complete. The weapon is now a perfect mirror of their own strategy.")
    print(f"'{FILENAME}' has been created.")

if __name__ == "__main__":
    main()
