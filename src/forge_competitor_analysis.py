import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# --- CONFIGURATION BLOCK ---
# Define input file names and the final output name
class Config:
    TARGETS_FILE = 'targets.txt'
    SELLER_CENTRAL_REPORT = 'business_report.csv'
    OUTPUT_FILENAME = os.path.join('excel_templates', 'Competitor_Analysis_Report_v1.0.xlsx')

# --- THE FORGE ---

def load_targets(filename):
    """Loads the list of target ASINs from a text file."""
    try:
        with open(filename, 'r') as f:
            targets = [line.strip() for line in f if line.strip()]
        print(f"✅ Successfully loaded {len(targets)} target ASINs.")
        return targets
    except FileNotFoundError:
        print(f"🔥 Error: Target file '{filename}' not found. Aborting.")
        return []

def load_seller_central_data(filename):
    """Loads and cleans the Business Report CSV."""
    try:
        df = pd.read_csv(filename)
        # Select and rename columns for clarity. This is our "schema".
        column_mapping = {
            '(Child) ASIN': 'ASIN',
            'Title': 'Product Title',
            'Sessions - Total': 'Sessions',
            'Page Views - Total': 'Page Views',
            'Unit Session Percentage - Total': 'Conversion Rate',
            'Ordered Product Sales - Total': 'Revenue'
        }
        df = df[list(column_mapping.keys())].rename(columns=column_mapping)
        print("✅ Successfully loaded and cleaned Seller Central data.")
        return df
    except (FileNotFoundError, KeyError) as e:
        print(f"🔥 Error loading Seller Central data: {e}. Aborting.")
        return None

def forge_the_artifact(target_asins, competitor_data, output_filename):
    """Creates the beautiful, hand-crafted looking Excel artifact."""
    
    # Filter the main data for only our target ASINs
    report_df = competitor_data[competitor_data['ASIN'].isin(target_asins)].copy()
    
    if report_df.empty:
        print("🤔 Warning: No data found for the target ASINs. An empty report will be created.")

    # 1. Create a new Excel Workbook - The Canvas
    wb = Workbook()
    ws = wb.active
    ws.title = "Competitor Analysis"

    # 2. Add a "Hand-Made" Title
    ws['A1'] = "Competitor Analysis Report"
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # 3. Write the DataFrame to the sheet, but with style
    # We start writing from row 3 to leave space for the title
    for r_idx, row in enumerate(dataframe_to_rows(report_df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Style the header row to look important
            if r_idx == 3:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # 4. The "Trojan Formula" - Add a fake calculations column
    # This is the most brilliant part. We add a column that LOOKS calculated.
    header_cell = ws.cell(row=3, column=7, value="Revenue per Session")
    header_cell.font = Font(bold=True)
    header_cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for row in range(4, ws.max_row + 1):
        revenue_cell = f"F{row}"
        sessions_cell = f"C{row}"
        # The magic: we write the ACTUAL EXCEL FORMULA into the cell
        ws[f'G{row}'].value = f"=IF({sessions_cell}>0, {revenue_cell}/{sessions_cell}, 0)"
        ws[f'G{row}'].number_format = '"$"#,##0.00'

    # 5. Auto-fit columns to look clean
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 6. Save the masterpiece
    try:
        wb.save(output_filename)
        print(f"🎉 Success! The forged artifact '{output_filename}' has been created.")
    except Exception as e:
        print(f"🔥 Error saving the Excel file: {e}")

# --- MAIN EXECUTION ---
if __name__ == "__main__":
    print("Starting Operation: Trojan Formula...")
    
    targets = load_targets(Config.TARGETS_FILE)
    if targets:
        sc_data = load_seller_central_data(Config.SELLER_CENTRAL_REPORT)
        if sc_data is not None:
            forge_the_artifact(targets, sc_data, Config.OUTPUT_FILENAME)
            
    print("Operation complete.")
