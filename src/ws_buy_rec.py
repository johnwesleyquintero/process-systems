# ws_buy_rec.py
# Wholesale Buy Recommendation Engine
# Python-native Excel generator for BUY/KILL/Prioritize decisions
# Author: John Wesley Quintero / WesAI
#
# --- LEGACY REVERSE ENGINEERING NOTES ---
# 1. UI/UX: 
#    - Frozen Panes: Standardized to row 4 (headers) and col B (ASIN/Title) for BUY/RESEARCH/KEEPA.
#    - AutoFilters: Enabled on all header rows to allow quick filtering by ASIN, Profit, etc.
#    - Column Widths: Heuristic-based (Title=55, ASIN=16, Numbers=12-14) to replicate legacy "pretty" look.
# 2. Logic:
#    - BUY Tab: Core decision sheet using VLOOKUPs to KEEPA and RESEARCH.
#    - KEEPA Tab: Implements weighted BB/AMZ averages. (30d=1, 90d=3, 180d=1 weights).
#    - IP Qty: Added 'In Buy Sheet?' check to cross-reference inventory planning with current review.
#    - ORDER: Dynamically pulls ASIN/Qty from BUY tab where Order Qty > 0.
# 3. Improvements:
#    - Removed sheet protection for unrestricted editing.
#    - Dynamic ranges (A:A) for VLOOKUPs to prevent breakages on large datasets.
#    - NA() and IFERROR() handling for cleaner mathematical outputs.

import os
import logging
import json
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- LOGGING ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- CONFIG & CONSTANTS ---
class ExcelConstants:
    """Centralized constants for Excel generation."""
    DEFAULT_MAX_ROWS: int = 200
    PASSWORD: str = "wesai"
    OUTPUT_DIR: str = "excel_templates"
    OUTPUT_FILE: str = os.path.join(OUTPUT_DIR, "BUY_RECOMMENDATIONS.xlsx")
    
    # Header Fill Colors
    HEADER_COLOR: str = "FFD966"
    SUMMARY_LABEL_COLOR: str = "D9D9D9"
    SUMMARY_VALUE_COLOR: str = "F2F2F2"

os.makedirs(ExcelConstants.OUTPUT_DIR, exist_ok=True)
MAX_ROWS = ExcelConstants.DEFAULT_MAX_ROWS

@dataclass
class ColumnMapper:
    """Manages column indexing and letter mapping for a worksheet."""
    tab_name: str
    headers: List[str]
    _map: Dict[str, int] = field(default_factory=dict, init=False)
    _letter_map: Dict[str, str] = field(default_factory=dict, init=False)

    def __post_init__(self):
        self._map = {h: i + 1 for i, h in enumerate(self.headers)}
        self._letter_map = {h: get_column_letter(i + 1) for i, h in enumerate(self.headers)}

    def get_index(self, header: str) -> int:
        """Returns the 1-based index of a header."""
        if header not in self._map:
            logger.error(f"Header '{header}' not found in tab '{self.tab_name}'")
            raise KeyError(f"Header '{header}' not found in tab '{self.tab_name}'. Available: {list(self._map.keys())}")
        return self._map[header]

    def get_letter(self, header: str) -> str:
        """Returns the Excel column letter for a header."""
        if header not in self._letter_map:
            logger.error(f"Header '{header}' not found in tab '{self.tab_name}'")
            raise KeyError(f"Header '{header}' not found in tab '{self.tab_name}'. Available: {list(self._letter_map.keys())}")
        return self._letter_map[header]

# --- FORMULA BUILDERS ---
class FormulaBuilder:
    """Helper for constructing common Excel formulas."""
    @staticmethod
    def vlookup(lookup_value: str, table_range: str, col_index: int, exact: bool = True, default: Any = 0) -> str:
        """
        Constructs a VLOOKUP formula wrapped in IFERROR.
        
        Args:
            lookup_value: The cell reference or value to look up.
            table_range: The range containing the lookup table (e.g., 'Sheet!$A:$C').
            col_index: The column index in the table to return.
            exact: Whether to perform an exact match.
            default: The value to return if the lookup fails.
        """
        match_type = "FALSE" if exact else "TRUE"
        return f'=IFERROR(VLOOKUP({lookup_value}, {table_range}, {col_index}, {match_type}), {default})'

    @staticmethod
    def match(lookup_val: str, lookup_range: str, match_type: int = 0) -> str:
        """
        Constructs a MATCH formula.
        
        Args:
            lookup_val: The cell reference or value to find.
            lookup_range: The range to search within.
            match_type: 0 for exact, 1 for less than, -1 for greater than.
        """
        return f"MATCH({lookup_val}, {lookup_range}, {match_type})"

    @staticmethod
    def count_items_in_list(cell: str, delimiter: str = ",") -> str:
        """
        Constructs a formula to count items in a delimited string.
        Useful for counting variation ASINs or other comma-separated lists.
        """
        return f"IF(ISBLANK({cell}), 1, LEN({cell}) - LEN(SUBSTITUTE({cell}, \"{delimiter}\", \"\")) + 1)"

    @staticmethod
    def if_blank(cell: str, value_if_blank: Any, value_if_not_blank: Any) -> str:
        """Constructs an IF(ISBLANK(...)) formula."""
        return f"IF(ISBLANK({cell}), {value_if_blank}, {value_if_not_blank})"

    @staticmethod
    def weighted_avg(components: List[tuple]) -> str:
        """
        Builds a weighted average formula that handles blanks and zeros gracefully.
        
        Args:
            components: A list of (value_cell, weight_cell) tuples.
        """
        num = " + ".join([f"IFERROR({v}*{w},0)" for v, w in components])
        den = " + ".join([f"IF({v}>0,{w},0)" for v, _ in components])
        return f"IFERROR(({num})/({den}), 0)"

class VLookupBuilder:
    """Specialized builder for VLOOKUPs to a specific range."""
    def __init__(self, table_range: str, default_val: Any = 0):
        self.table_range = table_range
        self.default_val = default_val

    def build(self, lookup_cell: str, col_index: int, exact: bool = True, default_val: Optional[Any] = None) -> str:
        default = default_val if default_val is not None else self.default_val
        return FormulaBuilder.vlookup(lookup_cell, self.table_range, col_index, exact, default)

# --- STYLES ---
HEADER_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
SUMMARY_LABEL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SUMMARY_VALUE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAPPED_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN_WRAPPED = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Protection Styles
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)

# Conditional Formatting Styles
RED_FONT = Font(color="FF9C0006")
RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
GREEN_FONT = Font(color="FF006100")
GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
ORANGE_FONT = Font(color="FF9C6500")
ORANGE_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")

# --- VALIDATORS ---
class ConfigValidator:
    """Validates the tab configuration for consistency and completeness."""
    @staticmethod
    def validate(config: Dict[str, Any]):
        required_tabs = ["BUY", "AZInsight_Data", "KEEPA", "IP Qty"]
        for tab in required_tabs:
            if tab not in config:
                raise ValueError(f"Missing required tab configuration: {tab}")
            
            tab_cfg = config[tab]
            if "headers" not in tab_cfg:
                raise ValueError(f"Tab '{tab}' is missing 'headers'")
            if "header_row" not in tab_cfg:
                raise ValueError(f"Tab '{tab}' is missing 'header_row'")
            
            # Check for unique headers
            headers = tab_cfg["headers"]
            if len(headers) != len(set(headers)):
                duplicates = [h for h in headers if headers.count(h) > 1]
                raise ValueError(f"Duplicate headers found in tab '{tab}': {set(duplicates)}")

        # Check for specific required headers for core logic
        required_headers = {
            "BUY": ["ASIN", "Cost", "Sell Price", "Profit", "ROI"],
            "AZInsight_Data": ["ASIN", "Sales Rank", "Estimated Number of Sales"],
            "KEEPA": ["ASIN", "BB Avg", "AMZ Avg"],
            "IP Qty": ["Barcode", "In Buy Sheet?"]
        }
        
        for tab, headers in required_headers.items():
            for header in headers:
                if header not in config[tab]["headers"]:
                    raise ValueError(f"Required header '{header}' missing from tab '{tab}'")

        logger.info("Configuration validation successful.")

# --- CONFIG LOADING ---
def load_config(file_path: str, max_rows: int) -> Dict[str, Any]:
    """Loads and interpolates the tab configuration from a JSON file."""
    with open(file_path, 'r') as f:
        config_str = f.read()
    
    # Interpolate dynamic values
    config_str = config_str.replace("{{MAX_ROWS}}", str(max_rows))
    config_str = config_str.replace("{{MAX_ROWS_PLUS_3}}", str(max_rows + 3))
    
    return json.loads(config_str)

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")
TABS_CONFIG = load_config(CONFIG_PATH, ExcelConstants.DEFAULT_MAX_ROWS)
ConfigValidator.validate(TABS_CONFIG)

# --- Column Mapping Cache ---
MAPPERS: Dict[str, ColumnMapper] = {
    tab: ColumnMapper(tab, config["headers"]) 
    for tab, config in TABS_CONFIG.items() 
    if "headers" in config
}

def get_col(tab_name: str, header_name: str) -> str:
    """
    Returns the Excel column letter for a given header name within a specific tab.
    Uses the pre-initialized ColumnMapper objects.
    """
    if tab_name not in MAPPERS:
        raise ValueError(f"Tab '{tab_name}' not found in MAPPERS.")
    
    return MAPPERS[tab_name].get_letter(header_name)

# --- HELPER FUNCTIONS ---
def setup_sheet(ws, tab_config: Dict[str, Any]):
    """
    Sets up headers, formatting, filters, and frozen panes for a worksheet.
    """
    headers = tab_config["headers"]
    header_row = tab_config["header_row"]
    
    logger.info(f"Setting up sheet: {ws.title}")

    # 1. Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color=ExcelConstants.HEADER_COLOR, 
                                end_color=ExcelConstants.HEADER_COLOR, fill_type="solid")
        cell.font = BOLD_FONT
        cell.alignment = WRAPPED_CENTER_ALIGN
        
        # Apply Column Widths
        col_letter = get_column_letter(col_idx)
        h_lower = header.lower()
        if ws.title == "README":
            ws.column_dimensions[col_letter].width = 30 if col_idx == 1 else 100
        elif any(kw in h_lower for kw in ["asin", "code", "sku", "barcode"]):
            ws.column_dimensions[col_letter].width = 16
        elif any(kw in h_lower for kw in ["title", "description", "name"]):
            ws.column_dimensions[col_letter].width = 55
        elif any(kw in h_lower for kw in ["link", "url", "page"]):
            ws.column_dimensions[col_letter].width = 10
        elif any(kw in h_lower for kw in ["qty", "bsr", "rank", "count", "offers", "sales", "stock", "items", "case count"]):
            ws.column_dimensions[col_letter].width = 12
        elif any(kw in h_lower for kw in ["price", "cost", "amount", "profit", "proceeds", "fee", "subtotal", "total cost", "roi", "margin"]):
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 15

        # Apply default column formatting based on header name
        fmt = 'General'
        if any(kw in h_lower for kw in ["price", "cost", "amount", "profit", "proceeds", "fee", "subtotal", "total cost"]):
            fmt = '"$"#,##0.00'
        elif any(kw in h_lower for kw in ["roi", "margin", "%", "chng", "weight", "oos"]):
            fmt = '0.00%'
        elif any(kw in h_lower for kw in ["qty", "bsr", "rank", "count", "offers", "sales", "stock", "items", "case count"]):
            fmt = '#,##0'
        elif h_lower == "roi/profit":
            fmt = '"OK";;"Not OK"'
            
        for row_idx in range(header_row + 1, header_row + MAX_ROWS + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt

    # 2. Enable AutoFilter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + MAX_ROWS}"

    # 3. Cell Protection Strategy
    if ws.title != "README":
        input_keywords = ["asin", "title", "cost", "case pack", "extra", "product id", "qty", "sku", "barcode", "replenishment", "lead time"]
        for col_idx, header in enumerate(headers, 1):
            h_lower = header.lower()
            is_input = any(kw in h_lower for kw in input_keywords) and "avg" not in h_lower and "oos" not in h_lower
            protection = UNLOCKED if is_input else LOCKED
            for row_idx in range(header_row + 1, header_row + MAX_ROWS + 1):
                ws.cell(row=row_idx, column=col_idx).protection = protection

    # 4. Freeze Panes
    if ws.title in ["BUY", "AZInsight_Data", "KEEPA"]:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    elif ws.title == "IP Qty":
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # 5. Summary Labels and Values
    if "summary_labels" in tab_config:
        for cell_ref, label in tab_config["summary_labels"].items():
            cell = ws[cell_ref]
            cell.value = label
            cell.fill = PatternFill(start_color=ExcelConstants.SUMMARY_LABEL_COLOR, 
                                    end_color=ExcelConstants.SUMMARY_LABEL_COLOR, fill_type="solid")
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            
    if "summary_values" in tab_config:
        for cell_ref, val in tab_config["summary_values"].items():
            cell = ws[cell_ref]
            cell.value = val
            cell.fill = PatternFill(start_color=ExcelConstants.SUMMARY_VALUE_COLOR, 
                                    end_color=ExcelConstants.SUMMARY_VALUE_COLOR, fill_type="solid")
            cell.alignment = CENTER_ALIGN
            
            # Unlock specific input cells in summary
            if ws.title in ["BUY", "KEEPA"] and cell_ref in ["B2", "C2", "D2", "F2", "L2", "Q2", "R2", "S2", "T2"]:
                cell.protection = UNLOCKED
            else:
                cell.protection = LOCKED
                
            if isinstance(val, (int, float)):
                if 0 < val < 1:
                    cell.number_format = '0.00%'
                elif val > 1000:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '0.00'

    # 6. Enable Sheet Protection
    ws.protection.sheet = True
    ws.protection.password = ExcelConstants.PASSWORD
    ws.protection.autoFilter = False

# --- INIT WORKBOOK ---
wb = Workbook()
# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Create tabs
for tab_name, config in TABS_CONFIG.items():
    ws = wb.create_sheet(title=tab_name)
    setup_sheet(ws, config)
    
    # Fill README content
    if tab_name == "README":
        for i, (section, desc) in enumerate(config["content"], 2):
            ws[f"A{i}"] = section
            ws[f"B{i}"] = desc
            ws[f"A{i}"].font = BOLD_FONT if section.isupper() else Font(bold=False)
            ws[f"B{i}"].alignment = LEFT_ALIGN_WRAPPED
            if section.isupper():
                ws[f"A{i}"].fill = SUMMARY_LABEL_FILL
                ws[f"B{i}"].fill = SUMMARY_LABEL_FILL

# --- FORMULA REFERENCES ---
# (Cache initialized on first get_col call)

# --- BUY TAB FORMULA INJECTION ---
buy_ws = wb["BUY"]
header_row_buy = TABS_CONFIG["BUY"]["header_row"]
start_row = header_row_buy + 1

# Map BUY headers to columns
b_asin = get_col("BUY", 'ASIN')
b_amz_title = get_col("BUY", 'AMZ Title')
b_est_qty = get_col("BUY", 'Est. Qty')
b_fba = get_col("BUY", 'New FBA')
b_mfn = get_col("BUY", 'New MFN')
b_offers = get_col("BUY", 'Total Offers')
b_sell = get_col("BUY", 'Sell Price')
b_bsr = get_col("BUY", 'Current BSR')
b_referral = get_col("BUY", 'Referral %')
b_sh = get_col("BUY", 'S & H')
b_proceeds = get_col("BUY", 'Proceeds')
b_profit = get_col("BUY", 'Profit')
b_roi = get_col("BUY", 'ROI')
b_opp = get_col("BUY", 'Sales Opp.')
b_sugg_pk = get_col("BUY", 'Suggested Pk')
b_pk_qty = get_col("BUY", 'Pack Qty')
b_odr_pk = get_col("BUY", 'Odr Qty (Pk)')
b_item_code = get_col("BUY", 'Item Code')
b_desc = get_col("BUY", 'Description')
b_cost = get_col("BUY", 'Cost')
b_units = get_col("BUY", 'Order Qty (Unit)')
b_amount = get_col("BUY", 'Order Amount')
b_var_weight = get_col("BUY", 'Var Weight')
b_case_pack = get_col("BUY", 'Case Pack')
b_ip_qty = get_col("BUY", 'IP Qty')
b_stock = get_col("BUY", 'Stock')
b_on_order = get_col("BUY", 'On Order')
b_velocity = get_col("BUY", 'Velocity')
b_units_sold = get_col("BUY", 'Units Sold')
b_offers_90d = get_col("BUY", 'Offers 90d')
b_roi_profit = get_col("BUY", 'ROI/Profit')
b_visible = get_col("BUY", 'Visible')
b_var_opp = get_col("BUY", 'Var. Opp.')
b_use_qty = get_col("BUY", 'Use Qty')
b_bsr30 = get_col("BUY", 'BSR 30 Avg')
b_chng30 = get_col("BUY", '% Chng30')
b_bsr90 = get_col("BUY", 'BSR 90 Avg')
b_chng90 = get_col("BUY", '% Chng90')
b_avail = get_col("BUY", 'Avail. Qty')
b_amz_boss = get_col("BUY", 'AMZ Boss')
b_amz_oos = get_col("BUY", 'AMZ OOS')
b_bbox_oos = get_col("BUY", 'Bbox OOS')
b_brand = get_col("BUY", 'Brand')
b_keepa_link = get_col("BUY", 'Keepa Link')
b_amz_link = get_col("BUY", 'AMZ Product Page')
b_gated = get_col("BUY", 'Check Gated in SC')

# --- LOOKUP INDICES ---
res_mapper = MAPPERS["AZInsight_Data"]
r_sell_price_idx = res_mapper.get_index("Sell Price")
r_bsr_idx = res_mapper.get_index("Sales Rank")
r_cost_idx = res_mapper.get_index("Purchase Price")
r_opp_idx = res_mapper.get_index("Estimated Number of Sales")
r_brand_idx = res_mapper.get_index("Brand")
r_title_idx = res_mapper.get_index("Title")
r_fba_idx = res_mapper.get_index("Low New Fba Price")
r_mfn_idx = res_mapper.get_index("Low New Mfn Price")
r_offers_idx = res_mapper.get_index("Total Offers")
r_referral_idx = res_mapper.get_index("Referral Fee")
r_proceeds_idx = res_mapper.get_index("Seller Proceeds")
r_sh_idx = res_mapper.get_index("Inbound Shipping Estimate")
r_case_pk_idx = res_mapper.get_index("Case Pk")
r_item_num_idx = res_mapper.get_index("Item Number")
r_size_idx = res_mapper.get_index("Size")

kee_mapper = MAPPERS["KEEPA"]
k_bsr30_idx = kee_mapper.get_index("Sales Rank: 30 days avg.")
k_bsr90_idx = kee_mapper.get_index("Sales Rank: 90 days avg.")
k_offers90_idx = kee_mapper.get_index("New Offer Count: 90 days avg.")
k_amz_oos_idx = kee_mapper.get_index("Amazon out of stock percentage: 90 days OOS %")
k_bbox_oos_idx = kee_mapper.get_index("Buy Box out of stock percentage: 90 days OOS %")
k_baseline_idx = kee_mapper.get_index("Baseline")
k_referral_idx = kee_mapper.get_index("Referral Fee %")

ip_mapper = MAPPERS["IP Qty"]
ip_barcode_idx = ip_mapper.get_index("Barcode")
ip_to_order_idx = ip_mapper.get_index("To Order")
ip_stock_idx = ip_mapper.get_index("Stock")
ip_on_order_idx = ip_mapper.get_index("On Order")
ip_sales_idx = ip_mapper.get_index("Sales")
ip_velocity_idx = ip_mapper.get_index("Adjusted Sales Velocity/mo")

# --- FORMULA BUILDERS ---
vlookup_res = VLookupBuilder("AZInsight_Data!$A:$AS", 0)
vlookup_kee = VLookupBuilder("KEEPA!$A:$AW", 0)
vlookup_ip = VLookupBuilder("IP Qty!$D:$R", 0) # Barcode is column D

for row in range(start_row, start_row + MAX_ROWS):
    asin_ref = f"{b_asin}{row}"
    
    # AZInsight_Data Tab Lookups
    buy_ws[f"{b_est_qty}{row}"] = vlookup_res.build(asin_ref, r_opp_idx)
    buy_ws[f"{b_sell}{row}"] = vlookup_res.build(asin_ref, r_sell_price_idx)
    buy_ws[f"{b_bsr}{row}"] = vlookup_res.build(asin_ref, r_bsr_idx)
    buy_ws[f"{b_cost}{row}"] = vlookup_res.build(asin_ref, r_cost_idx)
    
    # AMZ Title Lookup from AZInsight_Data Tab
    buy_ws[f"{b_amz_title}{row}"] = vlookup_res.build(asin_ref, r_title_idx, default_val='""')

    # Financial & Offer Lookups from AZInsight_Data Tab
    buy_ws[f"{b_fba}{row}"] = vlookup_res.build(asin_ref, r_fba_idx)
    buy_ws[f"{b_mfn}{row}"] = vlookup_res.build(asin_ref, r_mfn_idx)
    buy_ws[f"{b_offers}{row}"] = vlookup_res.build(asin_ref, r_offers_idx)
    buy_ws[f"{b_proceeds}{row}"] = vlookup_res.build(asin_ref, r_proceeds_idx)
    buy_ws[f"{b_sh}{row}"] = vlookup_res.build(asin_ref, r_sh_idx)
    buy_ws[f"{b_sugg_pk}{row}"] = vlookup_res.build(asin_ref, r_case_pk_idx, default_val=1)
    buy_ws[f"{b_item_code}{row}"] = vlookup_res.build(asin_ref, r_item_num_idx, default_val='""')
    buy_ws[f"{b_case_pack}{row}"] = vlookup_res.build(asin_ref, r_case_pk_idx, default_val=1)
    buy_ws[f"{b_avail}{row}"] = vlookup_res.build(asin_ref, r_size_idx)
    buy_ws[f"{b_desc}{row}"] = vlookup_res.build(asin_ref, r_title_idx, default_val='""')

    # Referral % from KEEPA (more accurate for percentage)
    buy_ws[f"{b_referral}{row}"] = vlookup_kee.build(asin_ref, k_referral_idx)

    # Pack Qty Logic
    buy_ws[f"{b_pk_qty}{row}"] = f'=IF({b_sugg_pk}{row}>0, {b_sugg_pk}{row}, 1)'

    # Profit & ROI (Wrapped in IFERROR)
    # Legacy Profit: Proceeds - Cost - (S & H)
    buy_ws[f"{b_profit}{row}"] = f"=IFERROR({b_proceeds}{row}-{b_cost}{row}-{b_sh}{row}, 0)"
    buy_ws[f"{b_roi}{row}"] = f"=IFERROR(IF({b_cost}{row}=0, 0, {b_profit}{row}/{b_cost}{row}), 0)"

    # KEEPA Tab Lookups
    buy_ws[f"{b_bsr30}{row}"] = vlookup_kee.build(asin_ref, k_bsr30_idx)
    buy_ws[f"{b_bsr90}{row}"] = vlookup_kee.build(asin_ref, k_bsr90_idx)
    buy_ws[f"{b_offers_90d}{row}"] = vlookup_kee.build(asin_ref, k_offers90_idx)
    buy_ws[f"{b_amz_oos}{row}"] = vlookup_kee.build(asin_ref, k_amz_oos_idx)
    buy_ws[f"{b_bbox_oos}{row}"] = vlookup_kee.build(asin_ref, k_bbox_oos_idx)

    # IP Qty Tab Lookups (Using ASIN as lookup value per legacy logic)
    # Note: Barcode is column D (index 4), so offset from D is (target_idx - 4 + 1)
    ip_offset = ip_barcode_idx - 1
    buy_ws[f"{b_ip_qty}{row}"] = vlookup_ip.build(asin_ref, ip_to_order_idx - ip_offset)
    buy_ws[f"{b_stock}{row}"] = vlookup_ip.build(asin_ref, ip_stock_idx - ip_offset)
    buy_ws[f"{b_on_order}{row}"] = vlookup_ip.build(asin_ref, ip_on_order_idx - ip_offset)
    buy_ws[f"{b_velocity}{row}"] = vlookup_ip.build(asin_ref, ip_velocity_idx - ip_offset)
    buy_ws[f"{b_units_sold}{row}"] = vlookup_ip.build(asin_ref, ip_sales_idx - ip_offset)

    # Trends (Wrapped in IFERROR)
    buy_ws[f"{b_chng30}{row}"] = f"=IFERROR(IF({b_bsr30}{row}=0, 0, ({b_bsr30}{row}-{b_bsr}{row})/{b_bsr}{row}), 0)"
    buy_ws[f"{b_chng90}{row}"] = f"=IFERROR(IF({b_bsr90}{row}=0, 0, ({b_bsr90}{row}-{b_bsr}{row})/{b_bsr}{row}), 0)"

    # Sales Opp (Legacy: Est Qty * (Days Stock / 30))
    stock_days = "$V$2"
    buy_ws[f"{b_opp}{row}"] = f"=IFERROR({b_est_qty}{row}*({stock_days}/30), 0)"

    # Visible (Legacy: If ROI/Profit > 0, 1, 0)
    buy_ws[f"{b_visible}{row}"] = f"=IF({b_roi_profit}{row}>0, 1, 0)"

    # Var. Opp. (Legacy logic)
    var_opp_formula = f"IF(ISNUMBER({b_var_weight}{row}), IF({b_amz_boss}{row}, {b_opp}{row}*{b_var_weight}{row}*$AN$2, {b_opp}{row}*{b_var_weight}{row}), NA())"
    buy_ws[f"{b_var_opp}{row}"] = f"=IFERROR({var_opp_formula}, NA())"

    # Use Qty (Legacy logic)
    max_order = "$U$2"
    buy_ws[f"{b_use_qty}{row}"] = f"=IFERROR(MIN({b_var_opp}{row}, {max_order}), 0)"

    # Order Qty (Unit)
    buy_ws[f"{b_units}{row}"] = f"={b_use_qty}{row}"
    
    # Order Amount
    buy_ws[f"{b_amount}{row}"] = f"={b_cost}{row}*{b_units}{row}"
    
    # Odr Qty (Pk)
    buy_ws[f"{b_odr_pk}{row}"] = f"=IFERROR({b_units}{row}/{b_pk_qty}{row}, 0)"

    # Brand Lookup from AZInsight_Data Tab
    buy_ws[f"{b_brand}{row}"] = vlookup_res.build(asin_ref, r_brand_idx, default_val='""')

    # Hyperlinks
    buy_ws[f"{b_keepa_link}{row}"] = f'=IF(OR({asin_ref}="", {asin_ref}=0), "", HYPERLINK("https://keepa.com/#!product/1-"&{asin_ref}, "Keepa"))'
    buy_ws[f"{b_amz_link}{row}"] = f'=IF(OR({asin_ref}="", {asin_ref}=0), "", HYPERLINK("https://www.amazon.com/dp/"&{asin_ref}, "Amazon"))'
    buy_ws[f"{b_gated}{row}"] = f'=IF(OR({asin_ref}="", {asin_ref}=0), "", HYPERLINK("https://sellercentral.amazon.com/product-search/search?q="&{asin_ref}, "Check Gated"))'

    # AMZ Boss (Legacy logic)
    # AND(VLOOKUP(...)<=$AO$2, Est. Qty >= $AM$2)
    baseline_lookup = vlookup_kee.build(asin_ref, k_baseline_idx)
    buy_ws[f"{b_amz_boss}{row}"] = f'=IFERROR(AND({baseline_lookup}<=$AO$2, {b_opp}{row} >= $AM$2), FALSE)'

    # ROI/Profit (Legacy Complex Logic)
    # Tiers based on Summary Values
    tier_logic = f"((({b_roi}{row}>=$AD$2)*({b_profit}{row}>=$AH$2)) + (({b_roi}{row}<$AD$2)*({b_roi}{row}>=$AE$2)*({b_profit}{row}>=$AG$2)) + (({b_roi}{row}<$AE$2)*({b_roi}{row}>=$AC$2)*({b_profit}{row}>=0)))"
    bsr_logic = f"IF($AI$2=0, 1, {b_bsr30}{row}<=$AI$2) * IF($AK$2=0, 1, {b_bsr90}{row}<=$AK$2)"
    other_filters = f"AND({b_opp}{row}>=$D$1, {b_bsr}{row}<=$I$2, {b_sell}{row}>=$H$2)"
    buy_ws[f"{b_roi_profit}{row}"] = f"=IF({b_pk_qty}{row}>1, ({tier_logic}*2 * {bsr_logic} * {other_filters}), ({tier_logic} * {bsr_logic} * {other_filters}))"

# --- BUY TAB CONDITIONAL FORMATTING & VALIDATION ---
# 1. ASIN Duplicate highlight
buy_ws.conditional_formatting.add(
    f"{b_asin}{start_row}:{b_asin}{start_row + MAX_ROWS}",
    Rule(type="duplicateValues", dxf=DifferentialStyle(fill=RED_FILL, font=RED_FONT))
)

# 2. AMZ Title Keywords (Frozen, Cool Ship, Spray)
b_title_col = get_col("BUY", "AMZ Title")
for kw in ["Frozen", "Cool Ship", "Spray"]:
    # Use relative reference for CF formula as suggested
    formula = f'NOT(ISERROR(SEARCH("{kw}",{b_title_col}{start_row})))'
    buy_ws.conditional_formatting.add(
        f"{b_title_col}{start_row}:{b_title_col}{start_row + MAX_ROWS}",
        Rule(type="expression", formula=[formula], dxf=DifferentialStyle(font=RED_FONT))
    )

# 3. Stock Low highlight (< 3)
buy_ws.conditional_formatting.add(
    f"{b_stock}{start_row}:{b_stock}{start_row + MAX_ROWS}",
    Rule(type="cellIs", operator='lessThan', formula=['3'], dxf=DifferentialStyle(fill=RED_FILL, font=RED_FONT))
)

# 4. Data Validation for "Use Qty" (Apply to whole range as suggested)
dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
buy_ws.add_data_validation(dv)
dv.add(f"{b_amz_boss}{start_row}:{b_amz_boss}{start_row + MAX_ROWS}") # Fixed range and column

# --- KEEPA TAB FORMULA INJECTION ---
keepa_ws = wb["KEEPA"]
start_row_keepa = TABS_CONFIG["KEEPA"]["header_row"] + 1

# Weights for wtd avg
w30 = "$B$2"
w90 = "$C$2"
w180 = "$D$2"
amz_less = "$F$2"

# Map KEEPA headers to columns
keepa_mapper = MAPPERS["KEEPA"]
k_asin = keepa_mapper.get_letter("ASIN")
k_baseline = keepa_mapper.get_letter("Baseline")
k_bb_avg = keepa_mapper.get_letter("BB Avg")
k_amz_avg = keepa_mapper.get_letter("AMZ Avg")
k_wtd_avg = keepa_mapper.get_letter("wtd avg")
k_vars = keepa_mapper.get_letter("Variations")
k_rev_wt = keepa_mapper.get_letter("Review Wt")
k_conf = keepa_mapper.get_letter("Confidence")

# Data columns from KEEPA tab
k_bb_30 = keepa_mapper.get_letter("Buy Box: 30 days avg.")
k_bb_90 = keepa_mapper.get_letter("Buy Box: 90 days avg.")
k_bb_180 = keepa_mapper.get_letter("Buy Box: 180 days avg.")
k_bb_curr = keepa_mapper.get_letter("Buy Box: Current")

k_amz_30 = keepa_mapper.get_letter("Amazon: 30 days avg.")
k_amz_90 = keepa_mapper.get_letter("Amazon: 90 days avg.")
k_amz_180 = keepa_mapper.get_letter("Amazon: 180 days avg.")
k_amz_curr = keepa_mapper.get_letter("Amazon: Current")
k_amz_oos = keepa_mapper.get_letter("Amazon out of stock percentage: 90 days OOS %")

k_var_asins = keepa_mapper.get_letter("Variation ASINs")
k_parent = keepa_mapper.get_letter("Parent ASIN")
k_rev_fmt = keepa_mapper.get_letter("Reviews: Reviews - Format Specific")

for row in range(start_row_keepa, start_row_keepa + MAX_ROWS):
    # Baseline
    keepa_ws[f"{k_baseline}{row}"] = f'=IF(ISBLANK({k_amz_oos}{row}),0,IF({k_amz_oos}{row}>=0.65, "BB", "AMZ"))'
    
    # Robust Weighted Average logic
    bb_components = [(f"{k_bb_30}{row}", w30), (f"{k_bb_90}{row}", w90), (f"{k_bb_180}{row}", w180)]
    bb_avg_formula = FormulaBuilder.weighted_avg(bb_components)
    keepa_ws[f"{k_bb_avg}{row}"] = f"=IF(AND({k_bb_30}{row}<{k_bb_90}{row},{k_bb_30}{row}<{k_bb_180}{row}), MIN({k_bb_30}{row},{k_bb_curr}{row}), {bb_avg_formula})"
    
    amz_components = [(f"{k_amz_30}{row}", w30), (f"{k_amz_90}{row}", w90), (f"{k_amz_180}{row}", w180)]
    amz_avg_formula = FormulaBuilder.weighted_avg(amz_components)
    keepa_ws[f"{k_amz_avg}{row}"] = f"=IF(AND({k_amz_30}{row}<{k_amz_90}{row},{k_amz_30}{row}<{k_amz_180}{row}), MIN({k_amz_30}{row},{k_amz_curr}{row}), {amz_avg_formula})"
    
    # wtd avg
    keepa_ws[f"{k_wtd_avg}{row}"] = f'=IF({k_baseline}{row}="BB",{k_bb_avg}{row},MIN({k_bb_avg}{row},{k_amz_avg}{row}*(1-{amz_less})))'
    
    # Variations count
    keepa_ws[f"{k_vars}{row}"] = f'={FormulaBuilder.count_items_in_list(f"{k_var_asins}{row}")}'
    
    # Review Weight (simplified legacy logic)
    keepa_ws[f"{k_rev_wt}{row}"] = f'=IF({k_vars}{row}=1,1,IF(OR(ISBLANK({k_rev_fmt}{row}),{k_rev_fmt}{row}=0),"Manual",{k_rev_fmt}{row}/MAX(SUMIF($AK:$AK,{k_parent}{row},$AL:$AL),1)))'

    # Confidence
    keepa_ws[f"{k_conf}{row}"] = f"=IF({k_vars}{row}=1, 1, 0)"

# --- AZInsight_Data TAB CONDITIONAL FORMATTING ---
res_ws = wb["AZInsight_Data"]
res_mapper = MAPPERS["AZInsight_Data"]
start_row_res = TABS_CONFIG["AZInsight_Data"]["header_row"] + 1
r_rank = res_mapper.get_letter("Sales Rank")
r_sales = res_mapper.get_letter("Estimated Number of Sales")
r_profit = res_mapper.get_letter("Profit")
r_margin = res_mapper.get_letter("Margin")
r_roi = res_mapper.get_letter("ROI")

# 1. Sales Rank highlight (Orange if 100k-200k)
res_ws.conditional_formatting.add(
    f"{r_rank}{start_row_res}:{r_rank}{start_row_res + MAX_ROWS}",
    Rule(type="cellIs", operator='between', formula=['100001', '200000'], dxf=DifferentialStyle(fill=ORANGE_FILL, font=ORANGE_FONT))
)

# 2. Estimated Sales highlight (Orange if 21-59)
res_ws.conditional_formatting.add(
    f"{r_sales}{start_row_res}:{r_sales}{start_row_res + MAX_ROWS}",
    Rule(type="cellIs", operator='between', formula=['21', '59'], dxf=DifferentialStyle(fill=ORANGE_FILL, font=ORANGE_FONT))
)

# 3. Profit highlight (Orange if 0.01-3.99)
res_ws.conditional_formatting.add(
    f"{r_profit}{start_row_res}:{r_profit}{start_row_res + MAX_ROWS}",
    Rule(type="cellIs", operator='between', formula=['0.01', '3.99'], dxf=DifferentialStyle(fill=ORANGE_FILL, font=ORANGE_FONT))
)

# 4. Margin/ROI highlight (Orange if < 30%)
for col in [r_margin, r_roi]:
    res_ws.conditional_formatting.add(
        f"{col}{start_row_res}:{col}{start_row_res + MAX_ROWS}",
        Rule(type="cellIs", operator='between', formula=['0.00001', '0.29999'], dxf=DifferentialStyle(fill=ORANGE_FILL, font=ORANGE_FONT))
    )

# 5. Summary Formulas (E2, F2, G2, O2, P2)
# Profit Total
buy_ws["E2"] = f"=SUM({b_profit}{start_row}:{b_profit}{start_row + MAX_ROWS})"
# Overall ROI
buy_ws["F2"] = f"=IFERROR(E2/SUM({b_cost}{start_row}:{b_cost}{start_row + MAX_ROWS}), 0)"
# Total Sales (Estimated)
buy_ws["G2"] = f"=SUM({b_opp}{start_row}:{b_opp}{start_row + MAX_ROWS})"
# Total Units
buy_ws["O2"] = f"=SUM({b_units}{start_row}:{b_units}{start_row + MAX_ROWS})"
# Total Amount
buy_ws["P2"] = f"=SUM({b_amount}{start_row}:{b_amount}{start_row + MAX_ROWS})"

# --- IP Qty TAB FORMULA INJECTION ---
ip_ws = wb["IP Qty"]
ip_mapper = MAPPERS["IP Qty"]
start_row_ip = TABS_CONFIG["IP Qty"]["header_row"] + 1
ip_barcode = ip_mapper.get_letter("Barcode")
ip_in_buy = ip_mapper.get_letter("In Buy Sheet?")

# Get BUY tab ASIN column letter for MATCH
buy_asin_col = MAPPERS["BUY"].get_letter("ASIN")

for row in range(start_row_ip, start_row_ip + MAX_ROWS):
    # Check if Barcode is in BUY tab ASIN column
    match_formula = FormulaBuilder.match(f"{ip_barcode}{row}", f"BUY!${buy_asin_col}:${buy_asin_col}", 0)
    ip_ws[f"{ip_in_buy}{row}"] = f'=IF(ISERROR({match_formula}), "No", "YES")'

# --- ORDER TAB LOGIC (Pivot-style Summary) ---
order_ws = wb["ORDER"]

# We use Dynamic Array formulas to create a "Pivot Table" effect.
# This requires Excel 365 or 2021+. If they have older, they can still use it as a static list.

# 1. Unique Item Codes where Qty > 0
# Formula: =UNIQUE(FILTER(BuyData[Item Code], (BuyData[Order Qty (Unit)]>0)*(BuyData[Item Code]<>"")))
order_ws["A2"] = '=IFERROR(UNIQUE(FILTER(BuyData[Item Code], (BuyData[Order Qty (Unit)]>0)*(BuyData[Item Code]<>""))), "No Orders")'

# 2. Description (using SPILL reference A2#)
# Formula: =XLOOKUP(A2#, BuyData[Item Code], BuyData[Description], "")
order_ws["B2"] = '=IF(A2="No Orders", "", XLOOKUP(A2#, BuyData[Item Code], BuyData[Description], ""))'

# 3. Unit Cost (Average or first match)
# Formula: =XLOOKUP(A2#, BuyData[Item Code], BuyData[Cost], 0)
order_ws["C2"] = '=IF(A2="No Orders", "", XLOOKUP(A2#, BuyData[Item Code], BuyData[Cost], 0))'

# 4. Total Order Qty (SUMIFS)
# Formula: =SUMIFS(BuyData[Order Qty (Unit)], BuyData[Item Code], A2#)
order_ws["D2"] = '=IF(A2="No Orders", "", SUMIFS(BuyData[Order Qty (Unit)], BuyData[Item Code], A2#))'

# 5. Total Order Amount (SUMIFS)
# Formula: =SUMIFS(BuyData[Order Amount], BuyData[Item Code], A2#)
order_ws["E2"] = '=IF(A2="No Orders", "", SUMIFS(BuyData[Order Amount], BuyData[Item Code], A2#))'

# 7. Wrap BUY data in an Excel Table for structured references
buy_last_col = get_column_letter(len(TABS_CONFIG["BUY"]["headers"]))
buy_table_range = f"A{TABS_CONFIG['BUY']['header_row']}:{buy_last_col}{3 + MAX_ROWS}"
buy_table = Table(displayName="BuyData", ref=buy_table_range)

# Add a default style
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
buy_table.tableStyleInfo = style
buy_ws.add_table(buy_table)

    # --- SAVE WORKBOOK ---
try:
    wb.save(ExcelConstants.OUTPUT_FILE)
    logger.info(f"BUY recommendation workbook generated: {ExcelConstants.OUTPUT_FILE}")
except PermissionError:
    logger.error(f"Could not save to {ExcelConstants.OUTPUT_FILE}. File is open.")
except Exception as e:
    logger.exception(f"Unexpected error during save: {e}")
