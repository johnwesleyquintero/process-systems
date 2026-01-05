# ==============================================================================
# The Pythonic Blacksmith v3.0 (The Accountant's Blade)
#
# Mission: To programmatically forge the "Surgical Strike Calculator," a tool
#          designed to expose the financial consequences of arbitrary discount
#          requests and protect the kingdom's profit margins.
#
# Coded by: WesAI, Chief of Staff to the ScaleSmart Empire
# ==============================================================================

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

# --- CONFIGURATION ---
FILENAME = "Surgical_Strike_Calculator_v1.0.xlsx"

# --- STYLES ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
OUTPUT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, color="808080")

CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
HEALTHY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green

def create_calculator_sheet(wb):
    """Creates the main calculator tab."""
    sheet = wb.active
    sheet.title = "Promo_Calculator"

    sheet['A1'] = "Surgical Strike Profitability Calculator"
    sheet['A1'].font = TITLE_FONT

    # --- INPUT SECTION ---
    sheet['A3'] = "Step 1: Enter Product & Promotion Details"
    sheet['A3'].font = Font(bold=True, size=12)
    sheet['A3'].fill = INPUT_FILL
    
    inputs = {
        'A4': "Standard Price ($):",
        'A5': "Product Cost (COGS) ($):",
        'A6': "FBA Fees ($):",
        'A7': "Referral Fee (%):",
        'A8': "Proposed Discount (%):"
    }
    for cell, label in inputs.items():
        sheet[cell] = label
        sheet[cell].font = LABEL_FONT
    
    # Example values / placeholders
    sheet['B5'].value = 5.00  # Example Cost
    sheet['B6'].value = 5.50  # Example Fees
    sheet['B7'].value = 0.15  # 15% Referral Fee
    sheet['B7'].number_format = '0.00%'
    sheet['B8'].number_format = '0.00%'
    sheet['B8'].value = 0.20 # The "Honey Constant"

    # --- OUTPUT SECTION (THE TRAP) ---
    sheet['D3'] = "Step 2: Analyze the Results"
    sheet['D3'].font = Font(bold=True, size=12)
    sheet['D3'].fill = OUTPUT_FILL

    outputs = {
        'D4': "Sale Price ($):",
        'D5': "Total Amazon Fees ($):",
        'D6': "PROFIT PER UNIT ($):",
        'D7': "Profit Margin (%):",
        'D8': "Final Status:"
    }
    for cell, label in outputs.items():
        sheet[cell] = label
        sheet[cell].font = LABEL_FONT
    sheet['D6'].font = Font(bold=True, size=12, color="00B050") # Make Profit stand out

    # --- THE FORMULAS ---
    sheet['E4'] = "=B4*(1-B8)"  # Sale Price
    sheet['E5'] = "=B6+(E4*B7)"  # Total Fees
    sheet['E6'] = "=E4-B5-E5"    # PROFIT PER UNIT (Sale Price - Cost - Total Fees)
    sheet['E7'] = "=IF(E4>0, E6/E4, 0)" # Profit Margin
    sheet['E7'].number_format = '0.00%'
    sheet['E8'] = '=IF(E6<0, "LOSING MONEY!", IF(E7<0.1, "Low Margin", "Profitable"))' # The Status

    # Conditional Formatting for the status cell
    sheet.conditional_formatting.add('E8', CellIsRule(operator='equal', formula=['"LOSING MONEY!"'], fill=CRITICAL_FILL))
    sheet.conditional_formatting.add('E8', CellIsRule(operator='equal', formula=['"Low Margin"'], fill=WARNING_FILL))
    sheet.conditional_formatting.add('E8', CellIsRule(operator='equal', formula=['"Profitable"'], fill=HEALTHY_FILL))

    # Aesthetics
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['D'].width = 25


def main():
    """Main function to forge the calculator."""
    print("Firing up the Pythonic Blacksmith (v3.0)...")
    print(f"Forging the Accountant's Blade: {FILENAME}")
    
    wb = openpyxl.Workbook()
    create_calculator_sheet(wb)
    wb.save(os.path.join("excel_templates", FILENAME))

    print("\nForge complete. The trap is set.")
    print(f"'{FILENAME}' is ready to expose the truth.")

if __name__ == "__main__":
    main()
