import openpyxl

def inspect_legacy_ui(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        # AutoFilter
        if ws.auto_filter.ref:
            print(f"AutoFilter: {ws.auto_filter.ref}")
        else:
            print("AutoFilter: None")
            
        # Tables
        if hasattr(ws, '_tables') and ws._tables:
            for table in ws._tables:
                print(f"Table: {table.name}, Range: {table.ref}")
        
        # Frozen Panes
        if ws.freeze_panes:
            print(f"Freeze Panes: {ws.freeze_panes}")
        else:
            print("Freeze Panes: None")
            
        # Check first few rows for filter-like formulas or notes
        print("First 2 rows sample:")
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            print(row[:10])

if __name__ == "__main__":
    inspect_legacy_ui(r"c:\Users\johnw\process-systems\excel_templates\Buy Rec 5.0 Legacy.xlsx")
