import openpyxl

def inspect_buy_formulas(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb['BUY']
    
    headers = []
    for cell in ws[3]: # Headers are on row 3
        headers.append(cell.value)
    
    print(f"Headers found: {headers}")
    
    # Get formulas from row 4
    formulas = {}
    for col_idx, header in enumerate(headers, 1):
        if header:
            cell = ws.cell(row=4, column=col_idx)
            formulas[header] = cell.value
            
    for header, formula in formulas.items():
        print(f"{header}: {formula}")

if __name__ == "__main__":
    inspect_buy_formulas(r"c:\Users\johnw\process-systems\excel_templates\Buy Rec 5.0 Legacy.xlsx")
