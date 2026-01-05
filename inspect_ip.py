import openpyxl

def inspect_order(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    if "ORDER" in wb.sheetnames:
        ws = wb["ORDER"]
        print(f"\n--- Sheet: ORDER ---")
        for row in ws.iter_rows(min_row=1, max_row=5):
            print([cell.value for cell in row])
    else:
        print("ORDER tab not found")

if __name__ == "__main__":
    inspect_order(r"c:\Users\johnw\process-systems\excel_templates\Buy Rec 5.0 Legacy.xlsx")
